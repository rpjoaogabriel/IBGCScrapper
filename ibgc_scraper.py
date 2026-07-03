"""
IBGC Certificados Scraper
=========================
Extrai TODOS os dados da tabela DataTables (client-side) do IBGC de uma vez,
sem navegar pelas 32 páginas, e gera o Excel formatado.

Instalação (rode UMA VEZ antes de usar):
    pip install playwright openpyxl
    playwright install chromium

Uso:
    python ibgc_scraper.py
"""

# ─── Categorias canônicas (ordem das colunas no Excel) ───────────────────────
CATS_ORDER = [
    "CCA",
    "CCA+ (Exame)",
    "CCA+ (Experiência)",
    "CCF (Exame)",
    "CCF (Experiência)",
    "CCF+",
    "CCoAud",
    "CCoAud+",
    "CGO",
    "CGO+",
]

# Mapa: valor exato do <option> → coluna canônica
CAT_MAP = {
    "CCA IBGC (Certificado para Conselheiro de Administração IBGC)":                            "CCA",
    "CCA+ IBGC (Exame) (Certificado para Conselheiro de Administração Experiente IBGC)":        "CCA+ (Exame)",
    "CCA+ IBGC (Experiência) (Certificado para Conselheiro de Administração Experiente IBGC)":  "CCA+ (Experiência)",
    "CCF IBGC (Exame) (Certificado para Conselheiro Fiscal IBGC)":                              "CCF (Exame)",
    "CCF IBGC (Experiência) (Certificado para Conselheiro Fiscal IBGC)":                        "CCF (Experiência)",
    "CCF+ IBGC (Certificado para Conselheiro Fiscal Experiente IBGC)":                          "CCF+",
    "CCoAud IBGC (Certificado para Membro de CoAud IBGC)":                                      "CCoAud",
    "CCoAud+ IBGC (Certificado para Membro de CoAud Experiente IBGC)":                          "CCoAud+",
    "CGO IBGC (Certificado para Governance Officer no IBGC)":                                   "CGO",
    "CGO+ IBGC (Experiencia) (Certificado para Governance Officer no IBGC)":                    "CGO+",
}

def normalize_cat(raw: str) -> str:
    raw = raw.strip()
    return CAT_MAP.get(raw, raw)


# ─── 1. Scraping via Playwright + DataTables API ──────────────────────────────

def scrape(url: str) -> list:
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        print(f"  Abrindo {url} ...")
        page.goto(url, wait_until="networkidle", timeout=60000)

        # Aguarda a tabela ser renderizada com ao menos 1 linha com dados
        page.wait_for_selector("#tableListaCCI tbody tr td", timeout=20000)

        # ── Estratégia principal: DataTables JS API ──────────────────────
        # A tabela é client-side: TODOS os dados já estão no DOM.
        # $.fn.dataTable.Api('#id').rows().data() retorna TUDO, ignorando paginação.
        raw = page.evaluate("""
            () => {
                // Tenta via DataTables API (mais limpo)
                try {
                    const api = new $.fn.dataTable.Api('#tableListaCCI');
                    return api.rows().data().toArray().map(r => [
                        (r[0] || '').replace(/<[^>]+>/g, '').trim(),
                        (r[1] || '').replace(/<[^>]+>/g, '').trim(),
                        (r[2] || '').trim(),
                    ]);
                } catch(e1) {}

                // Fallback: lê todas as <tr> do DOM (inclusive ocultas pela paginação)
                const rows = [];
                document.querySelectorAll('#tableListaCCI tbody tr').forEach(tr => {
                    const tds = tr.querySelectorAll('td');
                    if (tds.length >= 3) {
                        rows.push([
                            tds[0].innerText.trim(),
                            tds[1].innerText.trim(),
                            tds[2].innerText.trim(),
                        ]);
                    }
                });
                return rows;
            }
        """)

        browser.close()

    return raw


def parse_rows(raw: list) -> list[dict]:
    """Converte linhas brutas em lista de dicts com lista de categorias."""
    records = []
    for item in raw:
        if not isinstance(item, list) or len(item) < 3:
            continue
        nome, estado, cats_raw = str(item[0]).strip(), str(item[1]).strip(), str(item[2])

        # Categorias no 3º campo, separadas por <br> ou quebra de linha
        cats_text = cats_raw.replace("<br>", "\n").replace("<br/>", "\n").replace("<BR>", "\n")
        cats = []
        for line in cats_text.splitlines():
            line = line.strip()
            if line:
                cats.append(normalize_cat(line))

        if nome:
            records.append({"nome": nome, "estado": estado, "cats": cats})

    return records


def aggregate(records: list[dict]) -> list[dict]:
    """Agrupa por nome (case-insensitive) acumulando categorias e preservando estado."""
    merged: dict[str, dict] = {}
    for r in records:
        key = r["nome"].upper()
        if key not in merged:
            merged[key] = {"nome": r["nome"], "estado": r["estado"], "cats": set()}
        merged[key]["cats"].update(r["cats"])
        if r["estado"] and not merged[key]["estado"]:
            merged[key]["estado"] = r["estado"]

    return sorted(merged.values(), key=lambda x: x["nome"])


# ─── 2. Geração do Excel ──────────────────────────────────────────────────────

def build_excel(persons: list[dict], cats_found: list[str], output_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    H_FILL   = PatternFill("solid", fgColor="1F4E79")
    H_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    CHK_FILL = PatternFill("solid", fgColor="E2EFDA")
    CHK_FONT = Font(name="Arial", bold=True, color="375623", size=10)
    ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
    BODY     = Font(name="Arial", size=10)
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left",   vertical="center")
    border   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def apply_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill      = H_FILL
            cell.font      = H_FONT
            cell.alignment = CENTER
            cell.border    = border
        ws.row_dimensions[1].height = 22

    def write_tab(ws, col_headers: list[str], rows: list[dict], cat_cols: list[str]):
        ws.append(col_headers)
        apply_header(ws, len(col_headers))

        for i, p in enumerate(rows, start=2):
            row_vals = [p["nome"], p["estado"]] + [
                "✓" if c in p["cats"] else "" for c in cat_cols
            ]
            ws.append(row_vals)
            alt = (i % 2 == 0)

            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=i, column=ci)
                cell.border = border
                if ci <= 2:
                    cell.font      = BODY
                    cell.alignment = LEFT
                    if alt:
                        cell.fill = ALT_FILL
                else:
                    cell.alignment = CENTER
                    if val == "✓":
                        cell.fill = CHK_FILL
                        cell.font = CHK_FONT
                    elif alt:
                        cell.fill = ALT_FILL
                        cell.font = BODY
                    else:
                        cell.font = BODY

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 20
        for idx in range(3, 3 + len(cat_cols)):
            ws.column_dimensions[get_column_letter(idx)].width = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = openpyxl.Workbook()

    # Aba Geral
    ws_all = wb.active
    ws_all.title = "IBGC Certificados"
    write_tab(ws_all, ["Nome", "Estado"] + cats_found, persons, cats_found)

    # Uma aba por categoria (nome da aba ≤ 31 chars — limite do Excel)
    for cat in cats_found:
        subset = [p for p in persons if cat in p["cats"]]
        if not subset:
            continue
        ws_cat = wb.create_sheet(title=cat[:31])
        write_tab(ws_cat, ["Nome", "Estado"], subset, [])

    wb.save(output_path)

    print(f"\n✅  Arquivo salvo : {output_path}")
    print(f"    Profissionais  : {len(persons)}")
    print(f"    Categorias     : {', '.join(cats_found)}")
    print(f"    Abas geradas   : 'IBGC Certificados' + {len(cats_found)} abas de categoria")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    URL    = "https://www.ibgc.org.br/destaques/oficial-certificacao-certificados-ibgc"
    OUTPUT = "ibgc_certificados.xlsx"

    print("=" * 60)
    print("  IBGC Certificados Scraper")
    print("=" * 60)

    print("\n[1/3] Raspando dados (DataTables client-side) ...")
    try:
        raw = scrape(URL)
    except ImportError:
        print("\n❌  Playwright não instalado.")
        print("    Execute: pip install playwright && playwright install chromium")
        return
    except Exception as e:
        print(f"\n❌  Erro no scraping: {e}")
        raise

    if not raw:
        print("\n⚠️   Nenhum dado extraído.")
        print("    Verifique se o site exige login ou se a estrutura da página mudou.")
        return

    print(f"    {len(raw)} linhas brutas capturadas.")

    print("\n[2/3] Parseando e agregando registros ...")
    records = parse_rows(raw)
    persons = aggregate(records)
    print(f"    {len(persons)} profissionais únicos encontrados.")

    # Mantém ordem canônica; adiciona categorias desconhecidas ao final
    all_cats = {c for p in persons for c in p["cats"]}
    cats_found = [c for c in CATS_ORDER if c in all_cats]
    for c in sorted(all_cats - set(cats_found)):
        cats_found.append(c)

    print(f"    Categorias     : {', '.join(cats_found)}")

    print("\n[3/3] Gerando Excel ...")
    build_excel(persons, cats_found, OUTPUT)


if __name__ == "__main__":
    main()
